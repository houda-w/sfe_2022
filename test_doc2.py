from docx.shared import Cm
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Cm, Inches, Mm, Emu
import random
from openpyxl import load_workbook

#Import template document
template = DocxTemplate('pv_template2.docx')

#Import Excel data file
data_file = 'bd_test.xlsx'
wb = load_workbook(data_file)
sheet = wb['Sheet2']
max_row = sheet.max_row


#Load les races des genisses in a table 
elv_list = []
up_list = []
race_list = []
global_index = []

for cell_row in range( 2, max_row+1):
    nomElv = sheet.cell(row = cell_row, column = 4).value
    ref_up = sheet.cell(row = cell_row, column = 3).value
    race = sheet.cell(row = cell_row, column = 10).value
    race_list.append(race)
    #print(race)
    elv_list.append(nomElv)

    
unique_race_list = list(set(race_list))
unique_elv_list = list(set(elv_list))
print(race_list)


index_elv = []
temp_elv = []
for nomElv in unique_elv_list:
    
    for cell_row in range( 2, max_row+1):
        current_elv = sheet.cell(row = cell_row, column = 4).value
        if nomElv == current_elv:
            indice = [i for i, x in enumerate(elv_list) if x == current_elv]
        continue
           
    index_elv.append(indice)
    #print(index_elv)
#global_index.append(index_elv)
print(index_elv)
#print(global_index)

race_dict = {}
for nomElv in unique_elv_list:
    if nomElv not in race_dict:
        race_dict[nomElv] = []
        for cell_row in range( 2, max_row+1):
            current_elv = sheet.cell(row = cell_row, column = 4).value
            ind = elv_list.index(current_elv)
            if nomElv == current_elv:
               indice = [i for i, x in enumerate(elv_list) if x == current_elv]
               for  i in indice:
                   if i == ind:
                       race_dict[nomElv].append(race_list[i])
            continue
#print(race_dict)
   
            

#Populate table des génisses

#table_contents = []
for nomElv in unique_elv_list:
    for key in race_dict:
        if  nomElv == key:
            RACE = race_dict[key]
    context = {
    'table_contents': RACE,
    'Eleveur':nomElv,
    }
    ##print(table_contents)
    print(context) 
    #Declare template variables

    #Render automated report
    template.render(context)
    template.save(f'PV_{nomElv}.docx') #Creates Word doc and names it 
#print(table_contents)
#print(global_index)
#print(temp_elv)