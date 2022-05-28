import streamlit as st
import pandas as pd
import openpyxl 
from pathlib import Path
from docx.shared import Cm
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Cm, Inches, Mm, Emu
import random
from openpyxl import load_workbook
#from fpdf import FPDF
import os
import base64
import db_tryout
import download_zip
#from db_tryout import del_files


def get_binary_file_downloader_html(bin_file, file_label='File'):
    with open(bin_file, 'rb') as f:
        data = f.read()
    bin_str = base64.b64encode(data).decode()
    href = f'<a href="data:application/octet-stream;base64,{bin_str}" download="{os.path.basename(bin_file)}">Download {file_label}</a>'
    return href


def print_data(df,sheet_inp):
    #Import template document
    template = DocxTemplate('pv_template2.docx')
    wb = df
    sheet = wb[sheet_inp]
    max_row = sheet.max_row   

    #set working directory to save files
    base_dir = Path(__file__).parent
    output_dir = base_dir / "OUTPUT"
    

    #Load les races des genisses in a table 
    elv_list = []
    up_list = []
    race_list = []
    up_list = []
    zone_list = []
    global_index = []
    for cell_row in range( 2, max_row+1):
        nomElv = sheet.cell(row = cell_row, column = 4).value
        ref_up = sheet.cell(row = cell_row, column = 3).value
        race = sheet.cell(row = cell_row, column = 10).value
        zone = sheet.cell(row = cell_row, column = 1).value
        elv_list.append(nomElv)
        up_list.append(ref_up)
        race_list.append(race)
        zone_list.append(zone)
    unique_race_list = list(set(race_list))
    unique_elv_list = list(set(elv_list))
    
    
    
    #__________________fetching les races de chaque eleveurs et stocker dans un dict_________
    race_dict = {}
    up_dict = {}
    for nomElv in unique_elv_list:
        if nomElv not in race_dict:
           race_dict[nomElv] = []
           for cell_row in range( 2, max_row+1):
                current_elv = sheet.cell(row = cell_row, column = 4).value
                ind = elv_list.index(current_elv)
                if nomElv == current_elv:
                    indice = [i for i, x in enumerate(elv_list) if x == current_elv]
           for  i in indice:
               race_dict[nomElv].append(race_list[i])
        continue



    #__________________fetching les up de chaque eleveurs et stocker dans un dict_________
    for nomElv in unique_elv_list:
        if nomElv not in up_dict:
            up_dict[nomElv] = []
            for cell_row in range( 2, max_row+1):
                current_elv = sheet.cell(row = cell_row, column = 4).value
                ind = elv_list.index(current_elv)
                if nomElv == current_elv:
                    indice = [i for i, x in enumerate(elv_list) if x == current_elv]
                    for  i in indice:
                        if i == ind:
                           up_dict[nomElv].append(up_list[i])
                continue
        

    ##__________________fetching les zones de chaque eleveurs et stocker dans un dict_________
    zone_dict = {}
    for nomElv in unique_elv_list:
        if nomElv not in zone_dict:
            zone_dict[nomElv] = []
            for cell_row in range( 2, max_row+1):
                current_elv = sheet.cell(row = cell_row, column = 4).value
                ind = elv_list.index(current_elv)
                if nomElv == current_elv:
                    indice = [i for i, x in enumerate(elv_list) if x == current_elv]
                    for  i in indice:
                        if i == ind:
                           zone_dict[nomElv].append(zone_list[i])
                continue

    #Populate table des génisses

    context = {
      'table_contents': '',
      'Eleveur': '',
      'Ref_UP' : '',
      'Zone' : '',
      'Race' : '',
    }
    u= []
    #print(race_dict)
    for nomElv in unique_elv_list:
        for key in race_dict:  
            if  nomElv == key:
                RACE = race_dict[key]
                UP = up_dict[key] 
                ZONE = zone_dict[key]
                u = list(set(RACE))
        #continue
        context['table_contents']=RACE
        context['Eleveur'] = nomElv
        context['Ref_UP'] = UP[0] 
        context['Zone'] = ZONE[0]
        context['Race'] = u

        template.render(context)
        output_path = output_dir / f'PV_{nomElv}.docx'
        template.save(output_path)
        #insert pv doc in database
        #insertBLOB(nomElv,fic)
    #zip all PV files in a folder and download 
    download_zip.zip_folder()
        


def upl_data():
    #df2 = None
    st.subheader('impression des PV')
    st.caption('importation des données:')
    sheet_inp = st.text_input('Dpnner le nom de la feuille ')
    datafile = st.file_uploader("Upload file", type='xlsx')
    if datafile is not None:
        df = load_workbook(datafile)
        #sheet = pd.ExcelFile(datafile)
        #st.write(sheet)
        #df.to_excel(df2)
        #df.save()
        #print(df2)
        #st.table(df)
        #st.write(df)
        #st.write(df)
        print_data(df,sheet_inp)
    
def main():
    db_tryout.change_choice()
    #upl_data()

if __name__ == '__main__':
	  main()

   
