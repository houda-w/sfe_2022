import os
from docx.shared import Cm
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Cm, Inches, Mm, Emu
import random
from openpyxl import load_workbook
import win32com.client as win32  # pip install pywin32


#convert generated docs file to pdf 
def convert_to_pdf(doc):
    """Convert given word document to pdf"""
    word = win32.DispatchEx("Word.Application")
    new_name = doc.replace(".docx", r".pdf")
    worddoc = word.Documents.Open(doc)
    worddoc.SaveAs(new_name, FileFormat=17)
    worddoc.Close()
    return None



#Import template document
template = DocxTemplate('pv_template2.docx')

#Import Excel data file
data_file = 'bd_test.xlsx'
wb = load_workbook(data_file)
sheet = wb['Sheet2']
max_row = sheet.max_row


#Load les races des genisses in a table 
elv_list = []
up_list = []
race_list = []
up_list = []
zone_list = []
global_index = []

for cell_row in range( 2, max_row+1):
    nomElv = sheet.cell(row = cell_row, column = 4).value
    ref_up = sheet.cell(row = cell_row, column = 3).value
    race = sheet.cell(row = cell_row, column = 10).value
    zone = sheet.cell(row = cell_row, column = 1).value
    elv_list.append(nomElv)
    up_list.append(ref_up)
    race_list.append(race)
    zone_list.append(zone)
    
    #print(race)
unique_race_list = list(set(race_list))
unique_elv_list = list(set(elv_list))
print(race_list)


index_elv = []
temp_elv = []
for nomElv in unique_elv_list:
    
    for cell_row in range( 2, max_row+1):
        current_elv = sheet.cell(row = cell_row, column = 4).value
        if nomElv == current_elv:
            indice = [i for i, x in enumerate(elv_list) if x == current_elv]
        continue
           
    index_elv.append(indice)
    #print(index_elv)
#global_index.append(index_elv)
#print(index_elv)
#print(global_index)

#__________________fetching les races de chaque eleveurs et stocker dans un dict_________
race_dict = {}
up_dict = {}
for nomElv in unique_elv_list:
    if nomElv not in race_dict:
        race_dict[nomElv] = []
        for cell_row in range( 2, max_row+1):
            current_elv = sheet.cell(row = cell_row, column = 4).value
            #ind = elv_list.index(current_elv)
            #print(ind)
            if nomElv == current_elv:
               indice = [i for i, x in enumerate(elv_list) if x == current_elv]
               #print(indice)
        for  i in indice:
            race_dict[nomElv].append(race_list[i])
    continue
print(race_dict)


#__________________fetching les up de chaque eleveurs et stocker dans un dict_________
for nomElv in unique_elv_list:
    if nomElv not in up_dict:
        up_dict[nomElv] = []
        for cell_row in range( 2, max_row+1):
            current_elv = sheet.cell(row = cell_row, column = 4).value
            ind = elv_list.index(current_elv)
            if nomElv == current_elv:
               indice = [i for i, x in enumerate(elv_list) if x == current_elv]
               for  i in indice:
                   if i == ind:
                       up_dict[nomElv].append(up_list[i])
            continue
            
            
#__________________fetching les zones de chaque eleveurs et stocker dans un dict_________
zone_dict = {}
for nomElv in unique_elv_list:
    if nomElv not in zone_dict:
        zone_dict[nomElv] = []
        for cell_row in range( 2, max_row+1):
            current_elv = sheet.cell(row = cell_row, column = 4).value
            ind = elv_list.index(current_elv)
            if nomElv == current_elv:
               indice = [i for i, x in enumerate(elv_list) if x == current_elv]
               for  i in indice:
                   if i == ind:
                       zone_dict[nomElv].append(zone_list[i])
            continue


#Populate table des génisses

context = {
    'table_contents': '',
    'Eleveur': '',
    'Ref_UP' : '',
    'Zone' : '',
    'Race' : '',
}


#table_contents = []
u= []
for nomElv in unique_elv_list:
    for key in race_dict:
        if  nomElv == key:
            RACE = race_dict[key]
            UP = up_dict[key] 
            ZONE = zone_dict[key]
            u = list(set(RACE))
        #Declare template variables
    context['table_contents']=RACE
    context['Eleveur'] = nomElv
    context['Ref_UP'] = UP[0] 
    context['Zone'] = ZONE[0]
    context['Race'] = u
    
    output_name = f'PV_{nomElv}.docx'
     
    #print(context)
    #Render automated report
    #template.render(context)
    #template.save(output_name) #Creates Word doc and names it 

     # -- Convert to PDF [OPTIONAL]
    #path_to_word_document = os.path.join(os.getcwd(), output_name)
    #convert_to_pdf(path_to_word_document)
#print(table_contents)
#print(global_index)
#print(temp_elv)
