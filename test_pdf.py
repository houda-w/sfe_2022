from openpyxl import load_workbook #used to import the Excel Data
from datetime import datetime #used to work with date times
#used for merge tags. If there is an error, uninstall and install docx-mailmerge
from mailmerge import MailMerge


data_file = 'bd_test.xlsx'

wb = load_workbook(data_file)
sheet = wb['Sheet2']
max_row = sheet.max_row

elv_list = []
up_list = []
race_list = []
for cell_row in range( 3, max_row+1):
    nomElv = sheet.cell(row = cell_row, column = 4).value
    ref_up = sheet.cell(row = cell_row, column = 3).value
    race = sheet.cell(row = cell_row, column = 10).value
    race_list.append(race)
    elv_list.append(nomElv)
    up_list.append(ref_up)
unique_elv_list = list(set(elv_list)) #getting unique list of reps
unique_up_list = list(set(up_list))
unique_race_list = list(set(race_list))

# For each elv, create their order reports
for nomElv in unique_elv_list:
    list_genisse = []
    genisse_hist_list = []

    # Setting up Word document variables. Need to reuse template for each rep
    template_doc = "pv_template.docx"
    word_doc = MailMerge(template_doc)


    for cell_row in range(3 , max_row+1):
        #looping to check the current rep in spreadsheet
        current_elv = sheet.cell(row = cell_row, column = 4).value
        current_up =  sheet.cell(row = cell_row, column = 3).value

        #Checking to see if line item is for ccurr_elv
        if current_elv == nomElv:
            n_genisse = sheet.cell(row = cell_row, column = 6).value
            list_genisse.append(n_genisse)
            
            table_contents = []
            y = []
            for race in unique_race_list:
              t = sheet.cell(row = cell_row, column = 10).value
    
              table_contents.append({
                  'Value': t
                })
              y.append(t)

            #iterate through race listto merge it in the table 
            #for itrRace in list_race:
            #    current_race = sheet.cell(row = cell_row, column = 10).value
             #   tab_race = {
             #       'RACE' : current_race
             #   }


            #Appending bovs as a dict into a list, which will
            # be merged as a table

            elv_dict = {
            'Race' :  str(sheet.cell(row = cell_row, column = 10).value),
            'ZONE' : str(sheet.cell(row = cell_row, column = 1).value),
            'Ref_UP': current_up,
            'table_contents':table_contents
            }

            tab_dict = {
            "Designation" : 'GENISSES',
            "RACE" : str(sheet.cell(row = cell_row, column = 10).value)
            }
            #'N°_Genisse' : str(sheet.cell(row = cell_row, column = 6).value),

            #Appending dicts to merge as a table
            genisse_hist_list.append(tab_dict)
        

    # summing raw numbers into a total
    #total = sum(raw_subtotal_list)

    # Merging the name and formatting totals
    #print(list_race)
    #print([i for i in word_doc.get_merge_fields()])
    word_doc.merge(Eleveur=nomElv)
    #word_doc.merge_rows('Race',genisse_hist_list)
   # word_doc.merge(**tab_race)
    word_doc.render(**elv_dict)
    word_doc.write(f'PV_{nomElv}.docx') #Creates Word doc and names it

